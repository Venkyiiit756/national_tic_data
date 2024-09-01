import os
import requests
import json
import time
import datetime

# List of cities
cities = ['ahmedabad', 'ambala', 'amritsar', 'bengaluru', 'bathinda', 'bhopal', 'chandigarh', 'coimbatore', 'delhi-ncr', 'hyderabad']

#cities = ['ahmedabad', 'ambala', 'amritsar', 'bengaluru', 'bathinda', 'bhopal', 'chandigarh', 'coimbatore', 'delhi-ncr', 'gwalior', 'hubli', 'hyderabad', 'jaipur', 'kochi', 'kolkata', 'kota', 'lucknow', 'ludhiana', 'mangalore', 'mumbai', 'panipat', 'patna', 'pune', 'surat', 'thane', 'vadodara', 'vijayawada', 'bharuch', 'anand', 'panchkula', 'dhanbad', 'aurangabad', 'nashik', 'chennai', 'vizianagaram', 'navi-mumbai', 'nagpur', 'indore', 'goa', 'raipur', 'durgapur', 'faridabad', 'burdwan', 'vizag', 'kanpur', 'belagavi', 'siliguri', 'bhubaneswar', 'udaipur', 'kalyan', 'madurai', 'greater-noida', 'jalgaon', 'manipal', 'gurgaon', 'new-delhi', 'jodhpur', 'mysuru', 'kurnool', 'bhilwara', 'ajmer', 'gandhinagar', 'rajkot', 'bhiwadi', 'thrissur', 'jorhat', 'meerut', 'allahabad', 'balaghat', 'bhilai', 'bilaspur', 'bokaro', 'dehradun', 'gandhidham', 'guwahati', 'haridwar', 'jalandhar', 'jammu', 'kalaburagi', 'kolhapur', 'latur', 'mohali', 'moradabad', 'nanded', 'neemuch', 'pathankot', 'rudrapur', 'sikar', 'ujjain', 'bhimavaram', 'alappuzha', 'trichy', 'amalapuram', 'saharanpur', 'anakapalle', 'attili', 'kozhikode', 'chilakaluripet', 'chirala', 'draksharamam', 'eluru', 'guntakal', 'guntur', 'gurazala', 'kakinada', 'kasibugga', 'macherla', 'machilipatnam', 'mandapeta', 'mylavaram', 'nandyal', 'narasaraopet', 'narsapur', 'narsipatnam', 'nellore', 'nuziveedu', 'palasa', 'parvathipuram', 'payakaraopeta', 'peddapuram', 'rajahmundry', 'ramachandrapuram', 'ravulapalem', 'samalkota', 'sattenapalle', 'srikakulam', 'tadepalligudem', 'tekkali', 'tenali', 'tiruvuru', 'vinukonda', 'vissannapeta', 'rewari', 'vijayapura', 'gadag', 'kalpetta', 'kottarakara', 'malappuram', 'mananthavady', 'manjeri', 'mukkam', 'payyanur', 'perinthalmanna', 'ponnani', 'punalur', 'tirur', 'satna', 'puducherry', 'dharapuram', 'erode', 'hosur', 'karimangalam', 'katpadi', 'komarapalayam', 'pollachi', 'rajapalayam', 'sivakasi', 'thanjavur', 'theni', 'tirunelveli', 'tirupur', 'bhuvanagiri', 'dubbak', 'godavarikhani', 'kalwakurthy', 'karimnagar', 'khammam', 'madhira', 'mancherial', 'metpally', 'miryalaguda', 'nalgonda', 'nizamabad', 'parigi', 'sangareddi', 'sathupally', 'secunderabad', 'shadnagar', 'suryapet', 'tandur', 'warangal', 'bareilly', 'gorakhpur', 'jhansi', 'muzaffarnagar', 'kanhangad', 'avinashi', 'tirupati', 'pennagaram', 'kavali', 'alwar', 'dausa', 'mathura', 'shri-ganganagar', 'salem', 'junagadh', 'jagdalpur', 'jetpur', 'ankleshwar', 'balasore', 'rewa', 'ahmednagar', 'chandrapur', 'betul', 'guna', 'sarni', 'karnal', 'asansol', 'agra', 'aligarh', 'jamshedpur', 'andul', 'howrah', 'athagarh', 'roorkee', 'ropar', 'dhamtari', 'doraha', 'ghaziabad', 'ghazipur', 'himmatnagar', 'jalpaiguri', 'kaithal', 'noida', 'loni', 'kothapeta', 'narasannapeta', 'vuyyuru', 'vapi', 'krishnagiri', 'armoor', 'bhupalapalli', 'valsad', 'gondia', 'nadiad', 'kanchipuram', 'adipur', 'kawardha', 'kondagaon', 'indapur', 'rahuri', 'gajapathinagaram', 'palluruthy', 'akola', 'amravati', 'angamaly', 'bikaner', 'chhindwara', 'dindigul', 'durg', 'rajnandgaon', 'halol', 'jamnagar', 'khandwa', 'kishangarh', 'kollam', 'kurukshetra', 'muzaffarpur', 'malout', 'patan', 'ranchi', 'sangli', 'sonipat', 'ulhasnagar', 'vasai', 'yamunanagar', 'zirakpur', 'kotkapura', 'moga', 'yavatmal', 'dabhoi', 'dahod', 'baloda-bazar', 'etawah', 'haldwani', 'barwani', 'bulandshahr', 'hisar', 'kashipur', 'bhuj', 'bellary', 'dhuri', 'kalol', 'tinsukia', 'raigarh', 'bhavnagar', 'hathras', 'surendranagar', 'palghar', 'ichalkaranji', 'abohar', 'khambhat', 'baddi', 'edappal', 'cuddalore', 'hoshiarpur', 'dasuya', 'muvattupuzha', 'udgir', 'tadipatri', 'mudhol', 'ashtamichira', 'saligram', 'pattambi', 'khanapur', 'yellandu', 'chidambaram', 'chhatarpur', 'varanasi', 'mughalsarai', 'tirumalgiri', 'deesa', 'idar', 'nagaon', 'sanand', 'kichha', 'raebareli', 'navsari', 'mehsana', 'bayad', 'siddhpur', 'shillong', 'kodaly', 'podili', 'tuticorin', 'palakonda', 'thiruvarur', 'darsi', 'kozhinjampara', 'gudivada', 'ranebennur', 'cumbum', 'palakkad', 'chamarajanagara', 'nagercoil', 'varadium', 'gorantla', 'ulundurpet', 'vijapur', 'wani', 'kullu', 'bahadurgarh', 'digras', 'dhule', 'rourkela', 'pithampur', 'ashoknagar', 'mansa', 'kekri', 'cuttack', 'davanagere', 'shivpuri', 'gadarwara', 'bhusawal', 'kopargaon', 'kalol-panchmahal', 'pendra', 'dalli-rajhara', 'mangaldoi', 'sivasagar', 'firozepur', 'kodungallur', 'hanumangarh', 'karad', 'dewas', 'gaya', 'dimapur', 'itarsi', 'beawar', 'adoni', 'unnao', 'kareli', 'dharwad', 'chitradurga', 'jind', 'bagalkot', 'ratlam', 'sagwara', 'shivamogga', 'khanna', 'ambikapur', 'bhandara', 'faizabad', 'kundli', 'nimbahera', 'arambagh', 'aurangabad-west-bengal', 'krishnanagar', 'dubrajpur', 'thalassery', 'dhampur', 'khopoli', 'pandharpur', 'sitapur', 'dharamsala', 'sundargarh', 'jamkhed', 'addanki', 'ponnur', 'atmakur', 'anjar', 'bobbili', 'jhabua', 'nawanshahr', 'pithapuram', 'tatipaka', 'pipariya', 'seoni-malwa', 'hamirpur', 'korba', 'anantapur', 'gurdaspur', 'tumkur', 'banga', 'morena', 'azamgarh', 'pratapgarh-uttar-pradesh', 'dibrugarh', 'palampur', 'nawapara', 'fazilka', 'jangaon', 'rajam', 'jirapur', 'pilkhuwa', 'jalalabad', 'dehgam', 'bhadravati', 'port-blair', 'aruppukkottai', 'sivagangai', 'devakottai', 'nandurbar', 'palani', 'mannarkkad', 'repalle', 'berhampur', 'jiaganj', 'khargone', 'trivandrum', 'kanchikacherla', 'hanuman-junction', 'kaikaluru', 'naidupeta', 'atmakur-nellore', 'madanapalle', 'punganur', 'kuppam', 'jaggampeta', 'yeleswaram', 'hajipur', 'burhanpur', 'jajpur-road', 'kalimpong', 'bapatla', 'mukhed', 'motihari', 'tuni', 'baghapurana', 'lonavala', 'tanda', 'hindaun', 'singarayakonda', 'venkatagiri', 'dongargarh', 'bakhrahat', 'sultanpur', 'madugula', 'rupnagar', 'morinda', 'junnar', 'gopalganj', 'kolar', 'malikipuram', 'cooch-behar', 'shirpur', 'rayagada', 'mahabubabad', 'angul', 'mundra', 'nandigama', 'neelapalle', 'petlad', 'uthukottai', 'amreli', 'kunnamkulam', 'cheepurupalli', 'chittoor', 'dharmavaram', 'ganapavaram', 'kandukur', 'markapuram', 'mogalthur', 'nagari', 'nindra', 'proddatur', 'saluru', 'yemmiganur', 'adoor', 'amballur', 'chalakudy', 'guruvayur', 'irinjalakuda', 'kattanam', 'kodakara', 'koothattukulam', 'peringottukara', 'thalikulam', 'thiruvalla', 'valanchery', 'wadakkancherry', 'acharapakkam', 'ambur', 'ammaiyarkuppam', 'annur', 'anthiyur', 'arakkonam', 'batlagundu', 'eriyur', 'kallakurichi', 'kangeyam', 'karaikudi', 'kulithalai', 'kumbakonam', 'muthur', 'namakkal', 'perundalaiyur', 'perundurai', 'pudukkottai', 'ramanathapuram', 'sankarankovil', 'sathankulam', 'sethiyathope', 'somanur', 'tenkasi', 'thiruthani', 'thiruvannamalai', 'thuraiyur', 'tiruchendur', 'tirukoilur', 'udumalpet', 'vedaranyam', 'vellakoil', 'vellore', 'bonakal', 'chevella', 'nirmal', 'krosuru', 'cherukupalli', 'railway-koduru', 'godhra', 'dholka', 'tohana', 'giridih', 'bodi', 'villupuram', 'islampur', 'singrauli', 'sangrur', 'kotdwar', 'dharmaj', 'kapadvanj', 'fatehabad', 'punjai-puliampatti', 'north-paravur', 'kuchaman', 'kavindapadi', 'alangayam', 'agartala', 'mullanpur', 'palakollu', 'kadiri', 'kovvur', 'kadapa', 'chagallu', 'duggirala', 'allagadda', 'kotabommali', 'gudur-kurnool', 'palamaner', 'nallamada', 'kothacheruvu', 'nidadavolu', 'bestavaripeta', 'chebrolu', 'nellimarla', 'nallajerla', 'hiramandalam', 'palvancha', 'himatnagar', 'domkal', 'sultan-bathery', 'penuganchiprolu', 'pamur', 'rayachoti', 'tanguturu', 'uppada', 'pathapatnam', 'pedana', 'somandepalle', 'pamidi', 'shahpur', 'dandeli', 'sidlaghatta', 'udupi', 'koratagere', 'sirsi', 'chikkaballapura', 'tiptur', 'pavagada', 'ankola', 'kambainallur', 'tiruchengode', 'thammampatti', 'rasipuram', 'jalakandapuram', 'valapadi', 'tirupattur', 'kaveripattinam', 'mettur', 'chinnasalem', 'sankagiri', 'pallipalayam', 'bommidi', 'jammikunta', 'parkal', 'thorrur', 'chityal', 'kamanpur', 'choutuppal', 'huzurabad', 'huzurnagar', 'kothagudem', 'sirpur-kagaznagar', 'valigonda', 'ieeja', 'ambernath', 'akividu', 'bilgi', 'kodumur', 'panruti', 'rabkavi', 'daryapur', 'dhone', 'jami', 'ponduru', 'ichchapuram', 'satyavedu', 'gokavaram', 'deoghar', 'sindagi', 'katni', 'vijayamangalam', 'gingee', 'gobichettipalayam', 'idappadi', 'ambasamudram', 'kadayam', 'bellampally', 'mukerian', 'dinanagar', 'jayamkondan', 'banswada', 'kota-nellore', 'peravurani', 'mudalgi', 'sankeshwar', 'umbergaon', 'pandikkad', 'neyveli', 'jamkhambhaliya', 'katihar', 'bijainagar', 'kovilpatti', 'harur', 'sakti', 'hapur', 'nadia', 'sambalpur', 'jharsuguda', 'gauribidanur', 'sujangarh', 'puliyankudi', 'kurinjipadi', 'hardoi', 'silchar', 'mattanur', 'kotpad', 'thanipadi', 'uthangarai', 'patiala', 'pedakurapadu', 'chinnamanur', 'titagarh', 'patran', 'sirsa', 'batala', 'khamgaon', 'lakhimpur-uttar-pradesh', 'srikalahasti', 'sardulgarh', 'nagapattinam', 'mala', 'bhagalpur', 'washim', 'berachampa', 'changaramkulam', 'sambhal', 'jeypore', 'dabra', 'ponnamaravathi', 'thiruthuraipoondi', 'nambiyur', 'jamtara', 'narwana', 'sugauli', 'mandi-gobindgarh', 'paralakhemundi', 'arumbavur', 'sullurpeta', 'sirkali', 'koratla', 'bagepalli', 'kuzhithurai', 'banaganapalli', 'bagnan', 'jaunpur', 'mayiladuthurai', 'palacode', 'forbesganj', 'kotputli', 'pusad', 'morbi', 'narnaul', 'jejuri', 'khurja', 'raxaul', 'alangudi', 'ekma-chapra', 'bundu', 'barhi', 'alakode', 'dharmapuri', 'tezpur', 'silvassa', 'surandai', 'cherpulassery', 'golaghat', 'bijapur', 'veraval', 'kalyani', 'kalakad', 'karur', 'chiplun', 'paramathivelur', 'saharsa', 'purnia', 'keeranur', 'samastipur', 'dumka', 'lakhimpur-assam', 'nabadwip', 'kokrajhar', 'ramabhadrapuram', 'dhanera', 'attur', 'bhadohi', 'chanpatia', 'malda', 'raiganj', 'sumerpur', 'borsad', 'daman', 'karwar', 'nelakondapally', 'mandvi', 'khandela', 'sathyamangalam', 'periyakulam', 'bazpur', 'damoh', 'krishnarajanagara', 'siruvalur', 'talwandi-bhai', 'pileru', 'falna', 'pulivendula', 'ron', 'chotila', 'gundlupet', 'chennur', 'una', 'luxettipet', 'hoogly', 'supaul', 'hazaribagh', 'Nathdwara', 'tittagudi', 'Valliyur', 'Viralimalai', 'Balanagar', 'Koottanad', 'Pirangut', 'Shirur', 'marthandam', 'cherupuzha', 'budaun', 'undavalli', 'mau', 'akbarpur', 'kankipadu', 'kakkattil', 'arni', 'srivilliputhur', 'kollengode', 'velanthavalam', 'sindhanur', 'kheda', 'mummidivaram', 'aranthangi', 'nazirpur', 'gohana', 'raichur', 'jabalpur', 'nava-raipur', 'jagraon', 'cheruvathur', 'bharatpur', 'palasa-kasibugga', 'malkipuram', 'p-dharmavaram', 'chimakurthy', 'maduranthakam', 'sunam', 'waterland', 'budhlada', 'kottakkal', 'wandoor', 'kolathur', 'farrukhabad', 'nichlaul', 'vrindavan', 'edacheri', 'kangra', 'mandi', 'chamba', 'sirmaur', 'kinnaur', 'razole', 'nakrekal', 'firozabad', 'kurumassery', 'kondotty', 'vita', 'tumakuru', 'tindivanam', 'neyveli-township', 'bilaspur-himachal-pradesh', 'virudhachalam', 'nanjangud', 'bibinagar', 'bhongir', 'thavanampalle', 'umreth', 'bangarupallem', 'anaikatti', 'balasinor', 'chintamani', 'pali', 'kanipakam', 'nagireddypet', 'jalore', 'yelamanchili', 'elamanchili', 'panachamoodu', 'west-champaran', 'gudur-nellore', 'sathupalli', 'thiruppathur', 'kolargoldfields', 'hoshangabad', 'andimadam', 'gopiganj', 'ariyalur', 'parawada', 'basantpur', 'chandausi', 'ottapalam', 'falakata', 'kanigiri', 'mappedu', 'srinagar', 'nagaram', 'amangal', 'velangi', 'gudiyatham', 'cheyyar', 'arcot', 'ranipet', 'vaniyambadi', 'sonari', 'kondur', 'perambra', 'puttur', 'jhunjhunu', 'lalgudi', 'manapparai', 'krishnarajpet', 'malur', 'ranastalam', 'kothampakkam', 'rajampet', 'vyara', 'ghumarwin', 'melli', 'una-gujarat', 'kotma', 'mahalingpur', 'gauriganj', 'melattur', 'vandavasi', 'burhar', 'chaygaon', 'sankarapuram', 'sholinghur', 'halduchour', 'lalkuan', 'pantnagar', 'thiruvallur', 'marpally', 'kolappalur', 'gondal', 'bhainsa', 'kankroli', 'koheda', 'gadwal', 'haveri', 'pernambut', 'mettupalayam', 'sukma', 'dantewada', 'koilkuntla', 'ayodhya', 'marandahalli', 'karanja-lad', 'paonta-sahib', 'eral', 'authoor', 'putturap', 'gajuwaka', 'newtehri', 'kunnamangalam', 'ramgarh', 'hindupuram', 'vettaikaranpudur', 'kolumam', 'thiruvalangadu', 'koderma', 'devgad', 'handwara', 'baramulla', 'atraulia', 'kayamkulam', 'sathuvachari', 'sahjanwa', 'greater-mumbai', 'karanodai', 'palladam', 'penumuru', 'rath', 'pallipattu', 'pulgaon', 'curchorem', 'choondal', 'kalady', 'bengaluru', 'chennai', 'coimbatore', 'delhi-ncr', 'hyderabad', 'kolkata', 'madurai', 'puducherry', 'tirupur', 'trivandrum', 'vijayawada', 'vizag']

# Base API URL (replace 'hyderabad' with the actual city name in the loop)
base_url = "https://apiproxy.paytm.com/v3/movies/search/movie?meta=1&reqData=1&city={}&movieCode=o0gvbyvmx&version=3&site_id=6&channel=HTML5&child_site_id=370&client_id=ticketnew&clientId=ticketnew"


timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# Directory to save the JSON files
jsonDirectory = "cityJsons"+timestamp

# Create the directory if it doesn't exist
if not os.path.exists(jsonDirectory):
    os.makedirs(jsonDirectory)

# Loop through each city, fetch the JSON data, and save it
for city in cities:
    # Construct the API URL for the city
    api_url = base_url.format(city)
    
    # Fetch the JSON data from the API
    response = requests.get(api_url)
    

    
    # If the request was successful, save the JSON data
    if response.status_code == 200:
        json_data = response.json()
        
        # Save the JSON data to a file in the 'cityJsons' directory
        with open(os.path.join(jsonDirectory, f"{city}.json"), 'w') as json_file:
            json.dump(json_data, json_file, indent=4)
        
        print(f"Saved data for {city}")
    else:
        print(f"Failed to fetch data for {city}. Status code: {response.status_code}")

print("Data fetching complete.")


import json
import openpyxl
from datetime import datetime
import os
import zipfile



# Get all JSON file paths from the directory
json_file_paths = [os.path.join(jsonDirectory, file) for file in os.listdir(jsonDirectory) if file.endswith(".json")]

# Create a final Excel workbook to store combined data for all cities
final_workbook = openpyxl.Workbook()
final_sheet = final_workbook.active
final_sheet.title = "CombinedData"
final_headers = ["City", "Theater name", "AvailableTickets", "TotalTickets", "BookedTickets", "TotalGross", "BookedGross"]
final_sheet.append(final_headers)

# Accumulate data for each city across all theaters
city_data = {}

# Track generated city file paths
city_output_files = []

# Process each JSON file
for json_file_path in json_file_paths:
    # Load the JSON data from the file
    with open(json_file_path, "r") as file:
        data = json.load(file)

    # Extract the city name from the file name
    city_name = json_file_path.split('/')[-1].replace('.json', '')

    # Create a new Excel workbook for each city
    workbook = openpyxl.Workbook()
    sheet1 = workbook.active
    sheet1.title = "Sheet1"
    sheet2 = workbook.create_sheet(title="Sheet2")
    sheet3 = workbook.create_sheet(title="Sheet3")

    # Write headers for Sheet1
    sheet1_headers = ["Theater name", "Audi", "ShowTime", "Label", "sAvailTickets", "sTotalTickets", "sBookedTickets", "Price", "sTotalGross", "sBookedGross"]
    sheet1.append(sheet1_headers)

    # Write headers for Sheet2
    sheet2_headers = ["Theater name", "Audi", "Showtime", "AvailableTickets", "TotalTickets", "BookedTickets", "TotalGross", "BookedGross"]
    sheet2.append(sheet2_headers)

    # Write headers for Sheet3
    sheet3_headers = ["Theater name", "AvailableTickets", "TotalTickets", "BookedTickets", "TotalGross", "BookedGross"]
    sheet3.append(sheet3_headers)

    # Extract data from JSON
    sheet2_data = {}
    sheet3_data = {}

    for cinema in data["meta"]["cinemas"]:
        theater_name = cinema["name"]
        cinema_id = str(cinema["id"])
        sessions = data["pageData"]["sessions"].get(cinema_id, [])

        for session in sessions:
            audi = session["audi"]
            show_time = session["showTime"]
            key = (theater_name, audi, show_time)

            for area in session["areas"]:
                label = area["label"]
                sAvailTickets = area["sAvail"]
                sTotalTickets = area["sTotal"]
                price = area["price"]
                sBookedTickets = sTotalTickets - sAvailTickets

                sTotalGross = sTotalTickets * price
                sBookedGross = sBookedTickets * price

                # Write to Sheet1
                sheet1.append([theater_name, audi, show_time, label, sAvailTickets, sTotalTickets, sBookedTickets, price, sTotalGross, sBookedGross])

                # Accumulate data for Sheet2
                if key not in sheet2_data:
                    sheet2_data[key] = {
                        "AvailableTickets": 0,
                        "TotalTickets": 0,
                        "BookedTickets": 0,
                        "TotalGross": 0,
                        "BookedGross": 0
                    }
                sheet2_data[key]["AvailableTickets"] += sAvailTickets
                sheet2_data[key]["TotalTickets"] += sTotalTickets
                sheet2_data[key]["BookedTickets"] += sBookedTickets
                sheet2_data[key]["TotalGross"] += sTotalGross
                sheet2_data[key]["BookedGross"] += sBookedGross

                # Accumulate data for Sheet3 by theater (ignoring showtime)
                if theater_name not in sheet3_data:
                    sheet3_data[theater_name] = {
                        "AvailableTickets": 0,
                        "TotalTickets": 0,
                        "BookedTickets": 0,
                        "TotalGross": 0,
                        "BookedGross": 0
                    }
                sheet3_data[theater_name]["AvailableTickets"] += sAvailTickets
                sheet3_data[theater_name]["TotalTickets"] += sTotalTickets
                sheet3_data[theater_name]["BookedTickets"] += sBookedTickets
                sheet3_data[theater_name]["TotalGross"] += sTotalGross
                sheet3_data[theater_name]["BookedGross"] += sBookedGross

    # Write accumulated data to Sheet2
    for key, values in sheet2_data.items():
        theater_name, audi, show_time = key
        sheet2.append([theater_name, audi, show_time, values["AvailableTickets"], values["TotalTickets"], values["BookedTickets"], values["TotalGross"], values["BookedGross"]])

    # Write accumulated data to Sheet3
    for theater_name, values in sheet3_data.items():
        sheet3.append([theater_name, values["AvailableTickets"], values["TotalTickets"], values["BookedTickets"], values["TotalGross"], values["BookedGross"]])

        # Also add to the final combined sheet
        final_sheet.append([city_name, theater_name, values["AvailableTickets"], values["TotalTickets"], values["BookedTickets"], values["TotalGross"], values["BookedGross"]])

        # Accumulate data for the city
        if city_name not in city_data:
            city_data[city_name] = {
                "AvailableTickets": 0,
                "TotalTickets": 0,
                "BookedTickets": 0,
                "TotalGross": 0,
                "BookedGross": 0
            }
        city_data[city_name]["AvailableTickets"] += values["AvailableTickets"]
        city_data[city_name]["TotalTickets"] += values["TotalTickets"]
        city_data[city_name]["BookedTickets"] += values["BookedTickets"]
        city_data[city_name]["TotalGross"] += values["TotalGross"]
        city_data[city_name]["BookedGross"] += values["BookedGross"]

    # Fetch the current timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Save the workbook with the appropriate filename
    city_output_file_path = f"cityExcels/{city_name}_{timestamp}.xlsx"
    workbook.save(city_output_file_path)
    
    # Add the city output file path to the list
    city_output_files.append(city_output_file_path)

# Add a new sheet to the final workbook for consolidated data by city
consolidated_sheet = final_workbook.create_sheet(title="ConsolidatedCityData")
consolidated_headers = ["City", "AvailableTickets", "TotalTickets", "BookedTickets", "TotalGross", "BookedGross"]
consolidated_sheet.append(consolidated_headers)

# Write the consolidated data to the new sheet
for city_name, values in city_data.items():
    consolidated_sheet.append([city_name, values["AvailableTickets"], values["TotalTickets"], values["BookedTickets"], values["TotalGross"], values["BookedGross"]])

## Update the final output file path with the actual timestamp
final_output_file_path = f"cityExcels_{timestamp}/combined_cities_{timestamp}.xlsx"

# Create a directory to store the city Excel files
output_directory = f"cityExcels_{timestamp}/city_excel_data"
os.makedirs(output_directory, exist_ok=True)

# Save the final combined workbook with the updated sheet
final_workbook.save(final_output_file_path)

# Move the generated city files into the directory
for city_file in city_output_files:
    city_file_name = os.path.basename(city_file)
    os.rename(city_file, os.path.join(output_directory, city_file_name))

# Zip the directory
zip_file_path = f"cityExcels_{timestamp}/city_excel_data.zip"
with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zip_file:
    for root, dirs, files in os.walk(output_directory):
        for file in files:
            zip_file.write(os.path.join(root, file), os.path.relpath(os.path.join(root, file), output_directory))
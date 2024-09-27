import os
import requests
import json
import time
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import zipfile

# ----------------------------- Part 1: Fetching and Saving JSON Data -----------------------------

# List of cities (truncated for brevity)
cities = ['ahmedabad', 'ambala', 'amritsar', 'bengaluru', 'bathinda', 'bhopal', 'chandigarh', 'coimbatore', 'delhi-ncr', 'gwalior', 'hubli', 'hyderabad', 'jaipur', 'kochi', 'kolkata', 'kota', 'lucknow', 'ludhiana', 'mangalore', 'mumbai', 'panipat', 'patna', 'pune', 'surat', 'thane', 'vadodara', 'vijayawada', 'bharuch', 'anand', 'panchkula', 'dhanbad', 'aurangabad', 'nashik', 'chennai', 'vizianagaram', 'navi-mumbai', 'nagpur', 'indore', 'goa', 'raipur', 'durgapur', 'faridabad', 'burdwan', 'vizag', 'kanpur', 'belagavi', 'siliguri', 'bhubaneswar', 'udaipur', 'kalyan', 'madurai', 'greater-noida', 'jalgaon', 'manipal', 'gurgaon', 'new-delhi', 'jodhpur', 'mysuru', 'kurnool', 'bhilwara', 'ajmer', 'gandhinagar', 'rajkot', 'bhiwadi', 'thrissur', 'jorhat', 'meerut', 'allahabad', 'balaghat', 'bhilai', 'bilaspur', 'bokaro', 'dehradun', 'gandhidham', 'guwahati', 'haridwar', 'jalandhar', 'jammu', 'kalaburagi', 'kolhapur', 'latur', 'mohali', 'moradabad', 'nanded', 'neemuch', 'pathankot', 'rudrapur', 'sikar', 'ujjain', 'bhimavaram', 'alappuzha', 'trichy', 'amalapuram', 'saharanpur', 'anakapalle', 'attili', 'kozhikode', 'chilakaluripet', 'chirala', 'draksharamam', 'eluru', 'guntakal', 'guntur', 'gurazala', 'kakinada', 'kasibugga', 'macherla', 'machilipatnam', 'mandapeta', 'mylavaram', 'nandyal', 'narasaraopet', 'narsapur', 'narsipatnam', 'nellore', 'nuziveedu', 'palasa', 'parvathipuram', 'payakaraopeta', 'peddapuram', 'rajahmundry', 'ramachandrapuram', 'ravulapalem', 'samalkota', 'sattenapalle', 'srikakulam', 'tadepalligudem', 'tekkali', 'tenali', 'tiruvuru', 'vinukonda', 'vissannapeta', 'rewari', 'vijayapura', 'gadag', 'kalpetta', 'kottarakara', 'malappuram', 'mananthavady', 'manjeri', 'mukkam', 'payyanur', 'perinthalmanna', 'ponnani', 'punalur', 'tirur', 'satna', 'puducherry', 'dharapuram', 'erode', 'hosur', 'karimangalam', 'katpadi', 'komarapalayam', 'pollachi', 'rajapalayam', 'sivakasi', 'thanjavur', 'theni', 'tirunelveli', 'tirupur', 'bhuvanagiri', 'dubbak', 'godavarikhani', 'kalwakurthy', 'karimnagar', 'khammam', 'madhira', 'mancherial', 'metpally', 'miryalaguda', 'nalgonda', 'nizamabad', 'parigi', 'sangareddi', 'sathupally', 'secunderabad', 'shadnagar', 'suryapet', 'tandur', 'warangal', 'bareilly', 'gorakhpur', 'jhansi', 'muzaffarnagar', 'kanhangad', 'avinashi', 'tirupati', 'pennagaram', 'kavali', 'alwar', 'dausa', 'mathura', 'shri-ganganagar', 'salem', 'junagadh', 'jagdalpur', 'jetpur', 'ankleshwar', 'balasore', 'rewa', 'ahmednagar', 'chandrapur', 'betul', 'guna', 'sarni', 'karnal', 'asansol', 'agra', 'aligarh', 'jamshedpur', 'andul', 'howrah', 'athagarh', 'roorkee', 'ropar', 'dhamtari', 'doraha', 'ghaziabad', 'ghazipur', 'himmatnagar', 'jalpaiguri', 'kaithal', 'noida', 'loni', 'kothapeta', 'narasannapeta', 'vuyyuru', 'vapi', 'krishnagiri', 'armoor', 'bhupalapalli', 'valsad', 'gondia', 'nadiad', 'kanchipuram', 'adipur', 'kawardha', 'kondagaon', 'indapur', 'rahuri', 'gajapathinagaram', 'palluruthy', 'akola', 'amravati', 'angamaly', 'bikaner', 'chhindwara', 'dindigul', 'durg', 'rajnandgaon', 'halol', 'jamnagar', 'khandwa', 'kishangarh', 'kollam', 'kurukshetra', 'muzaffarpur', 'malout', 'patan', 'ranchi', 'sangli', 'sonipat', 'ulhasnagar', 'vasai', 'yamunanagar', 'zirakpur', 'kotkapura', 'moga', 'yavatmal', 'dabhoi', 'dahod', 'baloda-bazar', 'etawah', 'haldwani', 'barwani', 'bulandshahr', 'hisar', 'kashipur', 'bhuj', 'bellary', 'dhuri', 'kalol', 'tinsukia', 'raigarh', 'bhavnagar', 'hathras', 'surendranagar', 'palghar', 'ichalkaranji', 'abohar', 'khambhat', 'baddi', 'edappal', 'cuddalore', 'hoshiarpur', 'dasuya', 'muvattupuzha', 'udgir', 'tadipatri', 'mudhol', 'ashtamichira', 'saligram', 'pattambi', 'khanapur', 'yellandu', 'chidambaram', 'chhatarpur', 'varanasi', 'mughalsarai', 'tirumalgiri', 'deesa', 'idar', 'nagaon', 'sanand', 'kichha', 'raebareli', 'navsari', 'mehsana', 'bayad', 'siddhpur', 'shillong', 'kodaly', 'podili', 'tuticorin', 'palakonda', 'thiruvarur', 'darsi', 'kozhinjampara', 'gudivada', 'ranebennur', 'cumbum', 'palakkad', 'chamarajanagara', 'nagercoil', 'varadium', 'gorantla', 'ulundurpet', 'vijapur', 'wani', 'kullu', 'bahadurgarh', 'digras', 'dhule', 'rourkela', 'pithampur', 'ashoknagar', 'mansa', 'kekri', 'cuttack', 'davanagere', 'shivpuri', 'gadarwara', 'bhusawal', 'kopargaon', 'kalol-panchmahal', 'pendra', 'dalli-rajhara', 'mangaldoi', 'sivasagar', 'firozepur', 'kodungallur', 'hanumangarh', 'karad', 'dewas', 'gaya', 'dimapur', 'itarsi', 'beawar', 'adoni', 'unnao', 'kareli', 'dharwad', 'chitradurga', 'jind', 'bagalkot', 'ratlam', 'sagwara', 'shivamogga', 'khanna', 'ambikapur', 'bhandara', 'faizabad', 'kundli', 'nimbahera', 'arambagh', 'aurangabad-west-bengal', 'krishnanagar', 'dubrajpur', 'thalassery', 'dhampur', 'khopoli', 'pandharpur', 'sitapur', 'dharamsala', 'sundargarh', 'jamkhed', 'addanki', 'ponnur', 'atmakur', 'anjar', 'bobbili', 'jhabua', 'nawanshahr', 'pithapuram', 'tatipaka', 'pipariya', 'seoni-malwa', 'hamirpur', 'korba', 'anantapur', 'gurdaspur', 'tumkur', 'banga', 'morena', 'azamgarh', 'pratapgarh-uttar-pradesh', 'dibrugarh', 'palampur', 'nawapara', 'fazilka', 'jangaon', 'rajam', 'jirapur', 'pilkhuwa', 'jalalabad', 'dehgam', 'bhadravati', 'port-blair', 'aruppukkottai', 'sivagangai', 'devakottai', 'nandurbar', 'palani', 'mannarkkad', 'repalle', 'berhampur', 'jiaganj', 'khargone', 'trivandrum', 'kanchikacherla', 'hanuman-junction', 'kaikaluru', 'naidupeta', 'atmakur-nellore', 'madanapalle', 'punganur', 'kuppam', 'jaggampeta', 'yeleswaram', 'hajipur', 'burhanpur', 'jajpur-road', 'kalimpong', 'bapatla', 'mukhed', 'motihari', 'tuni', 'baghapurana', 'lonavala', 'tanda', 'hindaun', 'singarayakonda', 'venkatagiri', 'dongargarh', 'bakhrahat', 'sultanpur', 'madugula', 'rupnagar', 'morinda', 'junnar', 'gopalganj', 'kolar', 'malikipuram', 'cooch-behar', 'shirpur', 'rayagada', 'mahabubabad', 'angul', 'mundra', 'nandigama', 'neelapalle', 'petlad', 'uthukottai', 'amreli', 'kunnamkulam', 'cheepurupalli', 'chittoor', 'dharmavaram', 'ganapavaram', 'kandukur', 'markapuram', 'mogalthur', 'nagari', 'nindra', 'proddatur', 'saluru', 'yemmiganur', 'adoor', 'amballur', 'chalakudy', 'guruvayur', 'irinjalakuda', 'kattanam', 'kodakara', 'koothattukulam', 'peringottukara', 'thalikulam', 'thiruvalla', 'valanchery', 'wadakkancherry', 'acharapakkam', 'ambur', 'ammaiyarkuppam', 'annur', 'anthiyur', 'arakkonam', 'batlagundu', 'eriyur', 'kallakurichi', 'kangeyam', 'karaikudi', 'kulithalai', 'kumbakonam', 'muthur', 'namakkal', 'perundalaiyur', 'perundurai', 'pudukkottai', 'ramanathapuram', 'sankarankovil', 'sathankulam', 'sethiyathope', 'somanur', 'tenkasi', 'thiruthani', 'thiruvannamalai', 'thuraiyur', 'tiruchendur', 'tirukoilur', 'udumalpet', 'vedaranyam', 'vellakoil', 'vellore', 'bonakal', 'chevella', 'nirmal', 'krosuru', 'cherukupalli', 'railway-koduru', 'godhra', 'dholka', 'tohana', 'giridih', 'bodi', 'villupuram', 'islampur', 'singrauli', 'sangrur', 'kotdwar', 'dharmaj', 'kapadvanj', 'fatehabad', 'punjai-puliampatti', 'north-paravur', 'kuchaman', 'kavindapadi', 'alangayam', 'agartala', 'mullanpur', 'palakollu', 'kadiri', 'kovvur', 'kadapa', 'chagallu', 'duggirala', 'allagadda', 'kotabommali', 'gudur-kurnool', 'palamaner', 'nallamada', 'kothacheruvu', 'nidadavolu', 'bestavaripeta', 'chebrolu', 'nellimarla', 'nallajerla', 'hiramandalam', 'palvancha', 'himatnagar', 'domkal', 'sultan-bathery', 'penuganchiprolu', 'pamur', 'rayachoti', 'tanguturu', 'uppada', 'pathapatnam', 'pedana', 'somandepalle', 'pamidi', 'shahpur', 'dandeli', 'sidlaghatta', 'udupi', 'koratagere', 'sirsi', 'chikkaballapura', 'tiptur', 'pavagada', 'ankola', 'kambainallur', 'tiruchengode', 'thammampatti', 'rasipuram', 'jalakandapuram', 'valapadi', 'tirupattur', 'kaveripattinam', 'mettur', 'chinnasalem', 'sankagiri', 'pallipalayam', 'bommidi', 'jammikunta', 'parkal', 'thorrur', 'chityal', 'kamanpur', 'choutuppal', 'huzurabad', 'huzurnagar', 'kothagudem', 'sirpur-kagaznagar', 'valigonda', 'ieeja', 'ambernath', 'akividu', 'bilgi', 'kodumur', 'panruti', 'rabkavi', 'daryapur', 'dhone', 'jami', 'ponduru', 'ichchapuram', 'satyavedu', 'gokavaram', 'deoghar', 'sindagi', 'katni', 'vijayamangalam', 'gingee', 'gobichettipalayam', 'idappadi', 'ambasamudram', 'kadayam', 'bellampally', 'mukerian', 'dinanagar', 'jayamkondan', 'banswada', 'kota-nellore', 'peravurani', 'mudalgi', 'sankeshwar', 'umbergaon', 'pandikkad', 'neyveli', 'jamkhambhaliya', 'katihar', 'bijainagar', 'kovilpatti', 'harur', 'sakti', 'hapur', 'nadia', 'sambalpur', 'jharsuguda', 'gauribidanur', 'sujangarh', 'puliyankudi', 'kurinjipadi', 'hardoi', 'silchar', 'mattanur', 'kotpad', 'thanipadi', 'uthangarai', 'patiala', 'pedakurapadu', 'chinnamanur', 'titagarh', 'patran', 'sirsa', 'batala', 'khamgaon', 'lakhimpur-uttar-pradesh', 'srikalahasti', 'sardulgarh', 'nagapattinam', 'mala', 'bhagalpur', 'washim', 'berachampa', 'changaramkulam', 'sambhal', 'jeypore', 'dabra', 'ponnamaravathi', 'thiruthuraipoondi', 'nambiyur', 'jamtara', 'narwana', 'sugauli', 'mandi-gobindgarh', 'paralakhemundi', 'arumbavur', 'sullurpeta', 'sirkali', 'koratla', 'bagepalli', 'kuzhithurai', 'banaganapalli', 'bagnan', 'jaunpur', 'mayiladuthurai', 'palacode', 'forbesganj', 'kotputli', 'pusad', 'morbi', 'narnaul', 'jejuri', 'khurja', 'raxaul', 'alangudi', 'ekma-chapra', 'bundu', 'barhi', 'alakode', 'dharmapuri', 'tezpur', 'silvassa', 'surandai', 'cherpulassery', 'golaghat', 'bijapur', 'veraval', 'kalyani', 'kalakad', 'karur', 'chiplun', 'paramathivelur', 'saharsa', 'purnia', 'keeranur', 'samastipur', 'dumka', 'lakhimpur-assam', 'nabadwip', 'kokrajhar', 'ramabhadrapuram', 'dhanera', 'attur', 'bhadohi', 'chanpatia', 'malda', 'raiganj', 'sumerpur', 'borsad', 'daman', 'karwar', 'nelakondapally', 'mandvi', 'khandela', 'sathyamangalam', 'periyakulam', 'bazpur', 'damoh', 'krishnarajanagara', 'siruvalur', 'talwandi-bhai', 'pileru', 'falna', 'pulivendula', 'ron', 'chotila', 'gundlupet', 'chennur', 'una', 'luxettipet', 'hoogly', 'supaul', 'hazaribagh', 'Nathdwara', 'tittagudi', 'Valliyur', 'Viralimalai', 'Balanagar', 'Koottanad', 'Pirangut', 'Shirur', 'marthandam', 'cherupuzha', 'budaun', 'undavalli', 'mau', 'akbarpur', 'kankipadu', 'kakkattil', 'arni', 'srivilliputhur', 'kollengode', 'velanthavalam', 'sindhanur', 'kheda', 'mummidivaram', 'aranthangi', 'nazirpur', 'gohana', 'raichur', 'jabalpur', 'nava-raipur', 'jagraon', 'cheruvathur', 'bharatpur', 'palasa-kasibugga', 'malkipuram', 'p-dharmavaram', 'chimakurthy', 'maduranthakam', 'sunam', 'waterland', 'budhlada', 'kottakkal', 'wandoor', 'kolathur', 'farrukhabad', 'nichlaul', 'vrindavan', 'edacheri', 'kangra', 'mandi', 'chamba', 'sirmaur', 'kinnaur', 'razole', 'nakrekal', 'firozabad', 'kurumassery', 'kondotty', 'vita', 'tumakuru', 'tindivanam', 'neyveli-township', 'bilaspur-himachal-pradesh', 'virudhachalam', 'nanjangud', 'bibinagar', 'bhongir', 'thavanampalle', 'umreth', 'bangarupallem', 'anaikatti', 'balasinor', 'chintamani', 'pali', 'kanipakam', 'nagireddypet', 'jalore', 'yelamanchili', 'elamanchili', 'panachamoodu', 'west-champaran', 'gudur-nellore', 'sathupalli', 'thiruppathur', 'kolargoldfields', 'hoshangabad', 'andimadam', 'gopiganj', 'ariyalur', 'parawada', 'basantpur', 'chandausi', 'ottapalam', 'falakata', 'kanigiri', 'mappedu', 'srinagar', 'nagaram', 'amangal', 'velangi', 'gudiyatham', 'cheyyar', 'arcot', 'ranipet', 'vaniyambadi', 'sonari', 'kondur', 'perambra', 'puttur', 'jhunjhunu', 'lalgudi', 'manapparai', 'krishnarajpet', 'malur', 'ranastalam', 'kothampakkam', 'rajampet', 'vyara', 'ghumarwin', 'melli', 'una-gujarat', 'kotma', 'mahalingpur', 'gauriganj', 'melattur', 'vandavasi', 'burhar', 'chaygaon', 'sankarapuram', 'sholinghur', 'halduchour', 'lalkuan', 'pantnagar', 'thiruvallur', 'marpally', 'kolappalur', 'gondal', 'bhainsa', 'kankroli', 'koheda', 'gadwal', 'haveri', 'pernambut', 'mettupalayam', 'sukma', 'dantewada', 'koilkuntla', 'ayodhya', 'marandahalli', 'karanja-lad', 'paonta-sahib', 'eral', 'authoor', 'putturap', 'gajuwaka', 'newtehri', 'kunnamangalam', 'ramgarh', 'hindupuram', 'vettaikaranpudur', 'kolumam', 'thiruvalangadu', 'koderma', 'devgad', 'handwara', 'baramulla', 'atraulia', 'kayamkulam', 'sathuvachari', 'sahjanwa', 'greater-mumbai', 'karanodai', 'palladam', 'penumuru', 'rath', 'pallipattu', 'pulgaon', 'curchorem', 'choondal', 'kalady', 'bengaluru', 'chennai', 'coimbatore', 'delhi-ncr', 'hyderabad', 'kolkata', 'madurai', 'puducherry', 'tirupur', 'trivandrum', 'vijayawada', 'vizag']

# Base API URL (replace 'hyderabad' with the actual city name in the loop)
#base_url = "https://apiproxy.paytm.com/v3/movies/search/movie?meta=1&reqData=1&city={}&movieCode=ttdx_wuyn&version=3&site_id=6&channel=HTML5&child_site_id=370&client_id=ticketnew&clientId=ticketnew"
base_url = "https://apiproxy.paytm.com/v3/movies/search/movie?meta=1&reqData=1&city={}&movieCode=ttdx_wuyn&date=2024-09-28&version=3&site_id=6&channel=HTML5&child_site_id=370&client_id=ticketnew&clientId=ticketnew"

timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

# Directory to save the JSON files
json_directory = f"cityJsons_{timestamp}"

# Create the directory if it doesn't exist
os.makedirs(json_directory, exist_ok=True)

# Loop through each city, fetch the JSON data, and save it
for city in cities:
    # Construct the API URL for the city
    api_url = base_url.format(city)
    
    try:
        # Fetch the JSON data from the API
        response = requests.get(api_url, timeout=10)
        response.raise_for_status()  # Raise an error for bad status codes
    except requests.exceptions.RequestException as e:
        print(f"Error fetching data for {city}: {e}")
        continue  # Skip to the next city if there's an error
    
    # If the request was successful, save the JSON data
    if response.status_code == 200:
        json_data = response.json()
        
        # Save the JSON data to a file in the 'cityJsons' directory
        json_file_path = os.path.join(json_directory, f"{city}.json")
        with open(json_file_path, 'w', encoding='utf-8') as json_file:
            json.dump(json_data, json_file, indent=4)
        
        print(f"Saved data for {city}")
    else:
        print(f"Failed to fetch data for {city}. Status code: {response.status_code}")

print("Data fetching complete.")

# ----------------------------- Part 2: Processing JSON and Creating Excel Files -----------------------------

# Get all JSON file paths from the directory
json_file_paths = [os.path.join(json_directory, file) for file in os.listdir(json_directory) if file.endswith(".json")]

# Create a final Excel workbook to store combined data for all cities
final_workbook = openpyxl.Workbook()
final_sheet = final_workbook.active
final_sheet.title = "CombinedData"
final_headers = ["City", "Theater name", "AvailableTickets", "TotalTickets", "BookedTickets", "Occupancy (%)", "TotalGross", "BookedGross"]
final_sheet.append(final_headers)

# Apply header styling to final_sheet
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

for col_num, column_title in enumerate(final_headers, 1):
    cell = final_sheet.cell(row=1, column=col_num)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Set column widths for better readability
column_widths = [15, 25, 18, 15, 15, 15, 15, 15]
for i, width in enumerate(column_widths, 1):
    final_sheet.column_dimensions[get_column_letter(i)].width = width

# Accumulate data for each city across all theaters
city_data = {}

# Track generated city file paths
city_output_files = []

# Create a directory to store the city Excel files
city_excels_dir = f"cityExcels_{timestamp}/city_excel_data"
os.makedirs(city_excels_dir, exist_ok=True)

# Process each JSON file
for json_file_path in json_file_paths:
    # Load the JSON data from the file
    with open(json_file_path, "r", encoding='utf-8') as file:
        data = json.load(file)

    # Extract the city name from the file name
    city_name = os.path.splitext(os.path.basename(json_file_path))[0]

    # Create a new Excel workbook for each city
    workbook = openpyxl.Workbook()
    sheet1 = workbook.active
    sheet1.title = "DetailedData"
    sheet2 = workbook.create_sheet(title="AggregatedData_Per_Showtime")
    sheet3 = workbook.create_sheet(title="AggregatedData_Per_Theater")

    # Define headers for each sheet
    sheet1_headers = ["Theater name", "Audi", "ShowTime", "Label", "AvailableTickets", "TotalTickets", "BookedTickets", "Occupancy (%)", "Price", "TotalGross", "BookedGross"]
    sheet2_headers = ["Theater name", "Audi", "Showtime", "AvailableTickets", "TotalTickets", "BookedTickets", "Occupancy (%)", "TotalGross", "BookedGross"]
    sheet3_headers = ["Theater name", "AvailableTickets", "TotalTickets", "BookedTickets", "Occupancy (%)", "TotalGross", "BookedGross"]

    # Append headers
    sheet1.append(sheet1_headers)
    sheet2.append(sheet2_headers)
    sheet3.append(sheet3_headers)

    # Apply header styling to all sheets
    for sheet in [sheet1, sheet2, sheet3]:
        for col_num, column_title in enumerate(sheet1_headers if sheet == sheet1 else sheet2_headers if sheet == sheet2 else sheet3_headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Set column widths
            if sheet == sheet1:
                widths = [25, 10, 15, 10, 18, 15, 15, 15, 10, 15, 15]
            elif sheet == sheet2:
                widths = [25, 10, 15, 18, 15, 15, 15, 15, 15]
            else:
                widths = [25, 18, 15, 15, 15, 15, 15]
            sheet.column_dimensions[get_column_letter(col_num)].width = widths[col_num - 1]

    # Extract data from JSON
    sheet2_data = {}
    sheet3_data = {}

    for cinema in data.get("meta", {}).get("cinemas", []):
        theater_name = cinema.get("name", "Unknown Theater")
        cinema_id = str(cinema.get("id", "0"))
        sessions = data.get("pageData", {}).get("sessions", {}).get(cinema_id, [])

        for session in sessions:
            audi = session.get("audi", "Unknown Audi")
            show_time = session.get("showTime", "Unknown Time")
            key = (theater_name, audi, show_time)

            for area in session.get("areas", []):
                label = area.get("label", "N/A")
                sAvailTickets = area.get("sAvail", 0)
                sTotalTickets = area.get("sTotal", 0)
                price = area.get("price", 0.0)
                sBookedTickets = sTotalTickets - sAvailTickets

                sTotalGross = sTotalTickets * price
                sBookedGross = sBookedTickets * price

                # Calculate Occupancy
                occupancy = (sBookedTickets / sTotalTickets * 100) if sTotalTickets else 0

                # Write to Sheet1 (DetailedData)
                sheet1.append([
                    theater_name, audi, show_time, label, 
                    sAvailTickets, sTotalTickets, sBookedTickets, 
                    round(occupancy, 2), price, round(sTotalGross, 2), round(sBookedGross, 2)
                ])

                # Accumulate data for Sheet2 (AggregatedData_Per_Showtime)
                if key not in sheet2_data:
                    sheet2_data[key] = {
                        "AvailableTickets": 0,
                        "TotalTickets": 0,
                        "BookedTickets": 0,
                        "Occupancy": 0.0,
                        "TotalGross": 0.0,
                        "BookedGross": 0.0
                    }
                sheet2_data[key]["AvailableTickets"] += sAvailTickets
                sheet2_data[key]["TotalTickets"] += sTotalTickets
                sheet2_data[key]["BookedTickets"] += sBookedTickets
                sheet2_data[key]["TotalGross"] += sTotalGross
                sheet2_data[key]["BookedGross"] += sBookedGross

                # Calculate cumulative occupancy for Sheet2
                sheet2_data[key]["Occupancy"] = (sheet2_data[key]["BookedTickets"] / sheet2_data[key]["TotalTickets"] * 100) if sheet2_data[key]["TotalTickets"] else 0

                # Accumulate data for Sheet3 (AggregatedData_Per_Theater)
                if theater_name not in sheet3_data:
                    sheet3_data[theater_name] = {
                        "AvailableTickets": 0,
                        "TotalTickets": 0,
                        "BookedTickets": 0,
                        "Occupancy": 0.0,
                        "TotalGross": 0.0,
                        "BookedGross": 0.0
                    }
                sheet3_data[theater_name]["AvailableTickets"] += sAvailTickets
                sheet3_data[theater_name]["TotalTickets"] += sTotalTickets
                sheet3_data[theater_name]["BookedTickets"] += sBookedTickets
                sheet3_data[theater_name]["TotalGross"] += sTotalGross
                sheet3_data[theater_name]["BookedGross"] += sBookedGross

                # Calculate cumulative occupancy for Sheet3
                sheet3_data[theater_name]["Occupancy"] = (sheet3_data[theater_name]["BookedTickets"] / sheet3_data[theater_name]["TotalTickets"] * 100) if sheet3_data[theater_name]["TotalTickets"] else 0

    # Write accumulated data to Sheet2
    for key, values in sheet2_data.items():
        theater_name, audi, show_time = key
        sheet2.append([
            theater_name, audi, show_time, 
            values["AvailableTickets"], values["TotalTickets"], values["BookedTickets"], 
            round(values["Occupancy"], 2), round(values["TotalGross"], 2), round(values["BookedGross"], 2)
        ])

    # Write accumulated data to Sheet3
    for theater_name, values in sheet3_data.items():
        sheet3.append([
            theater_name, values["AvailableTickets"], values["TotalTickets"], 
            values["BookedTickets"], round(values["Occupancy"], 2), 
            round(values["TotalGross"], 2), round(values["BookedGross"], 2)
        ])

        # Also add to the final combined sheet
        final_sheet.append([
            city_name, theater_name, values["AvailableTickets"], values["TotalTickets"], 
            values["BookedTickets"], round(values["Occupancy"], 2), 
            round(values["TotalGross"], 2), round(values["BookedGross"], 2)
        ])

        # Accumulate data for the city
        if city_name not in city_data:
            city_data[city_name] = {
                "AvailableTickets": 0,
                "TotalTickets": 0,
                "BookedTickets": 0,
                "Occupancy": 0.0,
                "TotalGross": 0.0,
                "BookedGross": 0.0
            }
        city_data[city_name]["AvailableTickets"] += values["AvailableTickets"]
        city_data[city_name]["TotalTickets"] += values["TotalTickets"]
        city_data[city_name]["BookedTickets"] += values["BookedTickets"]
        city_data[city_name]["TotalGross"] += values["TotalGross"]
        city_data[city_name]["BookedGross"] += values["BookedGross"]
        # Update occupancy for city
        city_data[city_name]["Occupancy"] = (city_data[city_name]["BookedTickets"] / city_data[city_name]["TotalTickets"] * 100) if city_data[city_name]["TotalTickets"] else 0

    # Sort Sheet1 by Occupancy (column H, index 8)
    sheet1.auto_filter.ref = sheet1.dimensions
    data_rows = list(sheet1.iter_rows(min_row=2, values_only=True))
    data_rows_sorted = sorted(data_rows, key=lambda x: x[7] if x[7] is not None else 0)  # Occupancy is at index 7
    # Clear existing data and re-append sorted data
    sheet1.delete_rows(2, sheet1.max_row)
    for row in data_rows_sorted:
        sheet1.append(row)

    # Sort Sheet2 by Occupancy (column G, index 7)
    sheet2.auto_filter.ref = sheet2.dimensions
    data_rows = list(sheet2.iter_rows(min_row=2, values_only=True))
    data_rows_sorted = sorted(data_rows, key=lambda x: x[6] if x[6] is not None else 0)  # Occupancy is at index 6
    # Clear existing data and re-append sorted data
    sheet2.delete_rows(2, sheet2.max_row)
    for row in data_rows_sorted:
        sheet2.append(row)

    # Sort Sheet3 by Occupancy (column E, index 4)
    sheet3.auto_filter.ref = sheet3.dimensions
    data_rows = list(sheet3.iter_rows(min_row=2, values_only=True))
    data_rows_sorted = sorted(data_rows, key=lambda x: x[4] if x[4] is not None else 0)  # Occupancy is at index 4
    # Clear existing data and re-append sorted data
    sheet3.delete_rows(2, sheet3.max_row)
    for row in data_rows_sorted:
        sheet3.append(row)

    # Apply borders and alternating row colors for Sheet1, Sheet2, and Sheet3
    for sheet in [sheet1, sheet2, sheet3]:
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
            for cell in row:
                cell.border = thin_border
                # Apply alternating row colors
                if cell.row % 2 == 0:
                    cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
                else:
                    cell.fill = PatternFill(fill_type=None)

        # Format Occupancy column as percentage with two decimal places
        occupancy_col = 8 if sheet == sheet1 else 7 if sheet == sheet2 else 5
        for row in sheet.iter_rows(min_row=2, min_col=occupancy_col, max_col=occupancy_col):
            for cell in row:
                cell.number_format = '0.00%'

    # Write accumulated data to the final combined sheet has been done above

    # Save the workbook with the appropriate filename
    city_output_file_path = os.path.join(city_excels_dir, f"{city_name}_{timestamp}.xlsx")
    workbook.save(city_output_file_path)
    
    # Add the city output file path to the list
    city_output_files.append(city_output_file_path)

# Add a new sheet to the final workbook for consolidated data by city
consolidated_sheet = final_workbook.create_sheet(title="ConsolidatedCityData")
consolidated_headers = ["City", "AvailableTickets", "TotalTickets", "BookedTickets", "Occupancy (%)", "TotalGross", "BookedGross"]
consolidated_sheet.append(consolidated_headers)

# Apply header styling to consolidated_sheet
for col_num, column_title in enumerate(consolidated_headers, 1):
    cell = consolidated_sheet.cell(row=1, column=col_num)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center")
    # Set column widths
    widths = [15, 18, 15, 15, 15, 15, 15]
    consolidated_sheet.column_dimensions[get_column_letter(col_num)].width = widths[col_num - 1]

# Write the consolidated data to the new sheet
for city_name, values in city_data.items():
    consolidated_sheet.append([
        city_name, values["AvailableTickets"], values["TotalTickets"], 
        values["BookedTickets"], round(values["Occupancy"], 2), 
        round(values["TotalGross"], 2), round(values["BookedGross"], 2)
    ])

# Sort the consolidated sheet by Occupancy (column E, index 5)
consolidated_sheet.auto_filter.ref = consolidated_sheet.dimensions
data_rows = list(consolidated_sheet.iter_rows(min_row=2, values_only=True))
data_rows_sorted = sorted(data_rows, key=lambda x: x[4] if x[4] is not None else 0)  # Occupancy is at index 4
# Clear existing data and re-append sorted data
consolidated_sheet.delete_rows(2, consolidated_sheet.max_row)
for row in data_rows_sorted:
    consolidated_sheet.append(row)

# Apply borders and alternating row colors for consolidated_sheet
for row in consolidated_sheet.iter_rows(min_row=2, max_row=consolidated_sheet.max_row, max_col=consolidated_sheet.max_column):
    for cell in row:
        cell.border = thin_border
        # Apply alternating row colors
        if cell.row % 2 == 0:
            cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        else:
            cell.fill = PatternFill(fill_type=None)

# Format Occupancy column as percentage with two decimal places
for row in consolidated_sheet.iter_rows(min_row=2, min_col=5, max_col=5):
    for cell in row:
        cell.number_format = '0.00%'

# Fetch the current timestamp
current_timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

# Save the final combined workbook with the updated sheet
final_output_file_path = f"cityExcels_{timestamp}/combined_cities_{current_timestamp}.xlsx"
final_workbook.save(final_output_file_path)

# Move the generated city files into the directory has been done already

# Zip the directory
zip_file_path = f"cityExcels_{timestamp}/city_excel_data.zip"
with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zip_file:
    for root, dirs, files in os.walk(city_excels_dir):
        for file in files:
            zip_file.write(os.path.join(root, file), os.path.relpath(os.path.join(root, file), city_excels_dir))

print("Excel files created and zipped successfully.")

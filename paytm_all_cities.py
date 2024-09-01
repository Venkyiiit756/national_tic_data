import json
import openpyxl
from datetime import datetime
import os
import zipfile

# Define the directory containing the JSON files
directory = "cityJsons"

# Get all JSON file paths from the directory
json_file_paths = [os.path.join(directory, file) for file in os.listdir(directory) if file.endswith(".json")]

# Create a final Excel workbook to store combined data for all cities
final_workbook = openpyxl.Workbook()
final_sheet = final_workbook.active
final_sheet.title = "CombinedData"
final_headers = ["City", "Theater name", "AvailableTickets", "TotalTickets", "BookedTickets", "TotalGross", "BookedGross"]
final_sheet.append(final_headers)

# Accumulate data for each city across all theaters
city_data = {}

# Track generated city file paths
city_output_files = []

# Process each JSON file
for json_file_path in json_file_paths:
    # Load the JSON data from the file
    with open(json_file_path, "r") as file:
        data = json.load(file)

    # Extract the city name from the file name
    city_name = json_file_path.split('/')[-1].replace('.json', '')

    # Create a new Excel workbook for each city
    workbook = openpyxl.Workbook()
    sheet1 = workbook.active
    sheet1.title = "Sheet1"
    sheet2 = workbook.create_sheet(title="Sheet2")
    sheet3 = workbook.create_sheet(title="Sheet3")

    # Write headers for Sheet1
    sheet1_headers = ["Theater name", "Audi", "ShowTime", "Label", "sAvailTickets", "sTotalTickets", "sBookedTickets", "Price", "sTotalGross", "sBookedGross"]
    sheet1.append(sheet1_headers)

    # Write headers for Sheet2
    sheet2_headers = ["Theater name", "Audi", "Showtime", "AvailableTickets", "TotalTickets", "BookedTickets", "TotalGross", "BookedGross"]
    sheet2.append(sheet2_headers)

    # Write headers for Sheet3
    sheet3_headers = ["Theater name", "AvailableTickets", "TotalTickets", "BookedTickets", "TotalGross", "BookedGross"]
    sheet3.append(sheet3_headers)

    # Extract data from JSON
    sheet2_data = {}
    sheet3_data = {}

    for cinema in data["meta"]["cinemas"]:
        theater_name = cinema["name"]
        cinema_id = str(cinema["id"])
        sessions = data["pageData"]["sessions"].get(cinema_id, [])

        for session in sessions:
            audi = session["audi"]
            show_time = session["showTime"]
            key = (theater_name, audi, show_time)

            for area in session["areas"]:
                label = area["label"]
                sAvailTickets = area["sAvail"]
                sTotalTickets = area["sTotal"]
                price = area["price"]
                sBookedTickets = sTotalTickets - sAvailTickets

                sTotalGross = sTotalTickets * price
                sBookedGross = sBookedTickets * price

                # Write to Sheet1
                sheet1.append([theater_name, audi, show_time, label, sAvailTickets, sTotalTickets, sBookedTickets, price, sTotalGross, sBookedGross])

                # Accumulate data for Sheet2
                if key not in sheet2_data:
                    sheet2_data[key] = {
                        "AvailableTickets": 0,
                        "TotalTickets": 0,
                        "BookedTickets": 0,
                        "TotalGross": 0,
                        "BookedGross": 0
                    }
                sheet2_data[key]["AvailableTickets"] += sAvailTickets
                sheet2_data[key]["TotalTickets"] += sTotalTickets
                sheet2_data[key]["BookedTickets"] += sBookedTickets
                sheet2_data[key]["TotalGross"] += sTotalGross
                sheet2_data[key]["BookedGross"] += sBookedGross

                # Accumulate data for Sheet3 by theater (ignoring showtime)
                if theater_name not in sheet3_data:
                    sheet3_data[theater_name] = {
                        "AvailableTickets": 0,
                        "TotalTickets": 0,
                        "BookedTickets": 0,
                        "TotalGross": 0,
                        "BookedGross": 0
                    }
                sheet3_data[theater_name]["AvailableTickets"] += sAvailTickets
                sheet3_data[theater_name]["TotalTickets"] += sTotalTickets
                sheet3_data[theater_name]["BookedTickets"] += sBookedTickets
                sheet3_data[theater_name]["TotalGross"] += sTotalGross
                sheet3_data[theater_name]["BookedGross"] += sBookedGross

    # Write accumulated data to Sheet2
    for key, values in sheet2_data.items():
        theater_name, audi, show_time = key
        sheet2.append([theater_name, audi, show_time, values["AvailableTickets"], values["TotalTickets"], values["BookedTickets"], values["TotalGross"], values["BookedGross"]])

    # Write accumulated data to Sheet3
    for theater_name, values in sheet3_data.items():
        sheet3.append([theater_name, values["AvailableTickets"], values["TotalTickets"], values["BookedTickets"], values["TotalGross"], values["BookedGross"]])

        # Also add to the final combined sheet
        final_sheet.append([city_name, theater_name, values["AvailableTickets"], values["TotalTickets"], values["BookedTickets"], values["TotalGross"], values["BookedGross"]])

        # Accumulate data for the city
        if city_name not in city_data:
            city_data[city_name] = {
                "AvailableTickets": 0,
                "TotalTickets": 0,
                "BookedTickets": 0,
                "TotalGross": 0,
                "BookedGross": 0
            }
        city_data[city_name]["AvailableTickets"] += values["AvailableTickets"]
        city_data[city_name]["TotalTickets"] += values["TotalTickets"]
        city_data[city_name]["BookedTickets"] += values["BookedTickets"]
        city_data[city_name]["TotalGross"] += values["TotalGross"]
        city_data[city_name]["BookedGross"] += values["BookedGross"]

    # Fetch the current timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Save the workbook with the appropriate filename
    city_output_file_path = f"cityExcels/{city_name}_{timestamp}.xlsx"
    workbook.save(city_output_file_path)
    
    # Add the city output file path to the list
    city_output_files.append(city_output_file_path)

# Add a new sheet to the final workbook for consolidated data by city
consolidated_sheet = final_workbook.create_sheet(title="ConsolidatedCityData")
consolidated_headers = ["City", "AvailableTickets", "TotalTickets", "BookedTickets", "TotalGross", "BookedGross"]
consolidated_sheet.append(consolidated_headers)

# Write the consolidated data to the new sheet
for city_name, values in city_data.items():
    consolidated_sheet.append([city_name, values["AvailableTickets"], values["TotalTickets"], values["BookedTickets"], values["TotalGross"], values["BookedGross"]])

# Save the final combined workbook with the updated sheet
final_output_file_path = f"cityExcels/combined_cities_{timestamp}.xlsx"
final_workbook.save(final_output_file_path)

# Create a directory to store the city Excel files
output_directory = "cityExcels/city_excel_data"
os.makedirs(output_directory, exist_ok=True)

# Move the generated city files into the directory
for city_file in city_output_files:
    city_file_name = os.path.basename(city_file)
    os.rename(city_file, os.path.join(output_directory, city_file_name))

# Zip the directory
zip_file_path = "cityExcels/city_excel_data.zip"
with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zip_file:
    for root, dirs, files in os.walk(output_directory):
        for file in files:
            zip_file.write(os.path.join(root, file), os.path.relpath(os.path.join(root, file), output_directory))

# Return the path to the zip file
zip_file_path